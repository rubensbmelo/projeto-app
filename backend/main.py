from fastapi import FastAPI, HTTPException, Depends, status
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import StreamingResponse
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel, EmailStr
from typing import List, Optional
from bson import ObjectId
from datetime import datetime, timedelta, timezone
import os
import uuid
import logging
import io
from dotenv import load_dotenv
from jose import JWTError, jwt
import bcrypt
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ============================================================
# 1. Configurações
# ============================================================
load_dotenv()
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = FastAPI(title="ERP Vendas 2026 - Sistema Integrado")

# ============================================================
# 2. CORS
# ============================================================
ALLOWED_ORIGINS = os.getenv("ALLOWED_ORIGINS", "https://projeto-app-pink.vercel.app").split(",")

app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ============================================================
# 3. Conexão com MongoDB Atlas
# ============================================================
MONGODB_URL = os.getenv("MONGODB_URL", os.getenv("MONGO_URL", "mongodb://localhost:27017"))
DB_NAME = os.getenv("DB_NAME", "erp_database")
client = AsyncIOMotorClient(MONGODB_URL)
db = client[DB_NAME]

# ============================================================
# 4. Segurança — JWT + Bcrypt
# ============================================================
SECRET_KEY = os.getenv("SECRET_KEY", "TROQUE-NO-ENV")
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 480  # 8 horas

bearer_scheme = HTTPBearer()


def verificar_senha(senha_plain: str, senha_hash: str) -> bool:
    return bcrypt.checkpw(senha_plain.encode("utf-8"), senha_hash.encode("utf-8"))


def gerar_hash_senha(senha: str) -> str:
    return bcrypt.hashpw(senha.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")


def criar_token(data: dict) -> str:
    payload = data.copy()
    payload["exp"] = datetime.now(timezone.utc) + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    return jwt.encode(payload, SECRET_KEY, algorithm=ALGORITHM)


def verificar_token(credentials: HTTPAuthorizationCredentials = Depends(bearer_scheme)):
    try:
        payload = jwt.decode(credentials.credentials, SECRET_KEY, algorithms=[ALGORITHM])
        email: str = payload.get("sub")
        role: str = payload.get("role", "vendedor")
        if email is None:
            raise HTTPException(status_code=401, detail="Token inválido")
        return {"email": email, "role": role}
    except JWTError:
        raise HTTPException(status_code=401, detail="Token inválido ou expirado")


def apenas_admin(usuario=Depends(verificar_token)):
    if usuario.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Acesso restrito a administradores")
    return usuario


# ============================================================
# 5. Startup — cria admin padrão e índices
# ============================================================
@app.on_event("startup")
async def startup():
    # Índices para performance
    await db.usuarios.create_index("email", unique=True)
    await db.clientes.create_index("cnpj")
    await db.pedidos.create_index("numero_oc")
    await db.pedidos.create_index("cliente_id")
    await db.pedidos.create_index("status")
    # Índices orçamentos
    await db.orcamentos.create_index("numero_proposta")
    await db.orcamentos.create_index("cliente_id")
    await db.orcamentos.create_index("status")

    # Cria admin padrão se não existir
    admin_email = os.getenv("ADMIN_EMAIL", "rubensbmelo@hotmail.com")
    admin_nome = os.getenv("ADMIN_NOME", "Administrador")
    senha_env = os.getenv("ADMIN_PASSWORD", "TroqueEssaSenha@2026")

    if not await db.usuarios.find_one({"email": admin_email}):
        await db.usuarios.insert_one({
            "id": str(uuid.uuid4()),
            "nome": admin_nome,
            "email": admin_email,
            "senha_hash": gerar_hash_senha(senha_env),
            "role": "admin",
            "ativo": True,
            "criado_em": datetime.now(timezone.utc).isoformat()
        })
        logger.info(f"✅ Usuário admin criado: {admin_email}")


# ============================================================
# 6. Schemas
# ============================================================

class LoginSchema(BaseModel):
    email: str
    password: str


class UsuarioCreateSchema(BaseModel):
    nome: str
    email: EmailStr
    password: str
    role: Optional[str] = "vendedor"


class UsuarioUpdateSenhaSchema(BaseModel):
    senha_atual: str
    nova_senha: str


class MaterialSchema(BaseModel):
    nome: str
    numero_fe: Optional[str] = ""
    codigo: Optional[str] = ""
    descricao: Optional[str] = ""
    segmento: Optional[str] = ""
    preco_unit: float = 0.0
    peso_unit: float = 0.0
    comissao: float = 0.0


class ClienteSchema(BaseModel):
    nome: str
    cnpj: str
    endereco: Optional[str] = ""
    cidade: str
    estado: Optional[str] = ""
    comprador: Optional[str] = ""
    telefone: Optional[str] = ""
    email: Optional[str] = ""


class ItemPedido(BaseModel):
    material_id: str
    quantidade: int
    peso_calculado: float
    valor_unitario: Optional[float] = 0.0
    subtotal: Optional[float] = 0.0
    ipi: Optional[float] = 0.0
    comissao_percent: Optional[float] = 0.0
    comissao_valor: Optional[float] = 0.0


class PedidoSchema(BaseModel):
    cliente_id: Optional[str] = ""
    cliente_nome: str
    item_nome: Optional[str] = ""
    numero_fe: Optional[str] = ""
    itens: Optional[List[ItemPedido]] = []
    status: str = "PENDENTE"
    valor_total: float = 0.0
    peso_total: float = 0.0
    quantidade: Optional[int] = 0
    comissao_valor: Optional[float] = 0.0
    data_entrega: Optional[str] = ""
    numero_fabrica: Optional[str] = ""
    numero_oc: Optional[str] = ""
    condicao_pagamento: Optional[str] = ""
    observacoes: Optional[str] = ""
    pedido_mae: Optional[str] = ""
    historico_entrega: Optional[dict] = None


# ============================================================
# 7. Helpers
# ============================================================

def to_object_id(id_str: str) -> ObjectId:
    try:
        return ObjectId(id_str)
    except Exception:
        raise HTTPException(status_code=400, detail=f"ID inválido: {id_str}")


def serialize(doc: dict) -> dict:
    if "_id" in doc:
        doc["id"] = str(doc.pop("_id"))
    return doc


def gerar_numero_oc() -> str:
    return f"OC-{datetime.now().strftime('%Y%m%d%H%M%S')}"


# ============================================================
# 8. Autenticação
# ============================================================

@app.post("/api/auth/login")
async def login(data: LoginSchema):
    usuario = await db.usuarios.find_one({"email": data.email})
    if not usuario or not verificar_senha(data.password, usuario["senha_hash"]):
        raise HTTPException(status_code=401, detail="E-mail ou senha incorretos")
    if not usuario.get("ativo", True):
        raise HTTPException(status_code=403, detail="Usuário inativo. Contate o administrador.")

    token = criar_token({"sub": usuario["email"], "role": usuario.get("role", "vendedor")})
    return {
        "access_token": token,
        "token_type": "bearer",
        "user": {
            "id": usuario.get("id", ""),
            "nome": usuario["nome"],
            "email": usuario["email"],
            "role": usuario.get("role", "vendedor")
        }
    }


@app.get("/api/auth/me")
async def get_me(usuario=Depends(verificar_token)):
    doc = await db.usuarios.find_one({"email": usuario["email"]}, {"senha_hash": 0, "_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Usuário não encontrado")
    return doc


@app.put("/api/auth/trocar-senha")
async def trocar_senha(data: UsuarioUpdateSenhaSchema, usuario=Depends(verificar_token)):
    doc = await db.usuarios.find_one({"email": usuario["email"]})
    if not doc or not verificar_senha(data.senha_atual, doc["senha_hash"]):
        raise HTTPException(status_code=401, detail="Senha atual incorreta")
    await db.usuarios.update_one(
        {"email": usuario["email"]},
        {"$set": {"senha_hash": gerar_hash_senha(data.nova_senha)}}
    )
    return {"message": "Senha atualizada com sucesso!"}


# ============================================================
# 9. Gestão de Usuários (apenas admin)
# ============================================================

@app.get("/api/usuarios")
async def listar_usuarios(admin=Depends(apenas_admin)):
    usuarios = await db.usuarios.find({}, {"senha_hash": 0}).to_list(length=100)
    return [serialize(u) for u in usuarios]


@app.post("/api/usuarios", status_code=201)
async def criar_usuario(data: UsuarioCreateSchema, admin=Depends(apenas_admin)):
    if await db.usuarios.find_one({"email": data.email}):
        raise HTTPException(status_code=400, detail="E-mail já cadastrado")
    novo = {
        "id": str(uuid.uuid4()),
        "nome": data.nome,
        "email": data.email,
        "senha_hash": gerar_hash_senha(data.password),
        "role": data.role,
        "ativo": True,
        "criado_em": datetime.now(timezone.utc).isoformat()
    }
    await db.usuarios.insert_one(novo)
    return {"message": f"Usuário {data.email} criado!", "role": data.role}


@app.put("/api/usuarios/{usuario_id}/desativar")
async def desativar_usuario(usuario_id: str, admin=Depends(apenas_admin)):
    await db.usuarios.update_one({"id": usuario_id}, {"$set": {"ativo": False}})
    return {"message": "Usuário desativado!"}


@app.put("/api/usuarios/{usuario_id}/ativar")
async def ativar_usuario(usuario_id: str, admin=Depends(apenas_admin)):
    await db.usuarios.update_one({"id": usuario_id}, {"$set": {"ativo": True}})
    return {"message": "Usuário ativado!"}


@app.put("/api/usuarios/{usuario_id}")
async def editar_usuario(usuario_id: str, data: dict, admin=Depends(apenas_admin)):
    campos = {k: v for k, v in data.items() if k in ["nome", "email", "role"]}
    if not campos:
        raise HTTPException(status_code=400, detail="Nenhum campo válido para atualizar")
    await db.usuarios.update_one({"id": usuario_id}, {"$set": campos})
    return {"message": "Usuário atualizado!"}


@app.put("/api/usuarios/{usuario_id}/reset-senha")
async def resetar_senha(usuario_id: str, data: dict, admin=Depends(apenas_admin)):
    nova_senha = data.get("nova_senha")
    if not nova_senha or len(nova_senha) < 6:
        raise HTTPException(status_code=400, detail="Senha deve ter no mínimo 6 caracteres")
    await db.usuarios.update_one(
        {"id": usuario_id},
        {"$set": {"senha_hash": gerar_hash_senha(nova_senha)}}
    )
    return {"message": "Senha resetada com sucesso!"}


@app.delete("/api/usuarios/{usuario_id}")
async def deletar_usuario(usuario_id: str, admin=Depends(apenas_admin)):
    doc = await db.usuarios.find_one({"id": usuario_id})
    if not doc:
        raise HTTPException(status_code=404, detail="Usuário não encontrado")
    await db.usuarios.delete_one({"id": usuario_id})
    return {"message": "Usuário removido!"}


# ============================================================
# 10. Materiais
# ============================================================

@app.get("/api/materiais")
async def listar_materiais(usuario=Depends(verificar_token)):
    materiais = await db.materiais.find().to_list(length=500)
    return [serialize(m) for m in materiais]


@app.post("/api/materiais", status_code=201)
async def criar_material(material: MaterialSchema, usuario=Depends(verificar_token)):
    doc = material.dict()
    doc["criado_em"] = datetime.now(timezone.utc).isoformat()
    doc["criado_por"] = usuario["email"]
    result = await db.materiais.insert_one(doc)
    return {"id": str(result.inserted_id), "message": "Material salvo!"}


@app.put("/api/materiais/{material_id}")
async def atualizar_material(material_id: str, material: MaterialSchema, usuario=Depends(verificar_token)):
    doc = material.dict()
    doc["atualizado_em"] = datetime.now(timezone.utc).isoformat()
    doc["atualizado_por"] = usuario["email"]
    result = await db.materiais.update_one(
        {"_id": to_object_id(material_id)},
        {"$set": doc}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Material não encontrado")
    return {"message": "Material atualizado!"}


@app.delete("/api/materiais/{material_id}")
async def deletar_material(material_id: str, admin=Depends(apenas_admin)):
    result = await db.materiais.delete_one({"_id": to_object_id(material_id)})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Material não encontrado")
    return {"message": "Material removido!"}


# ============================================================
# 11. Clientes
# ============================================================

@app.get("/api/clientes")
async def listar_clientes(usuario=Depends(verificar_token)):
    clientes = await db.clientes.find().to_list(length=500)
    return [serialize(c) for c in clientes]


@app.get("/api/clientes/{cliente_id}")
async def buscar_cliente(cliente_id: str, usuario=Depends(verificar_token)):
    cliente = await db.clientes.find_one({"_id": to_object_id(cliente_id)})
    if not cliente:
        raise HTTPException(status_code=404, detail="Cliente não encontrado")
    return serialize(cliente)


@app.post("/api/clientes", status_code=201)
async def criar_cliente(cliente: ClienteSchema, usuario=Depends(verificar_token)):
    ultimo = await db.clientes.find_one({}, sort=[("criado_em", -1)])
    num = 1
    if ultimo and "referencia" in ultimo:
        try:
            num = int(ultimo["referencia"].split("-")[1]) + 1
        except Exception:
            pass
    doc = cliente.dict()
    doc["referencia"] = f"CLI-{num:04d}"
    doc["criado_em"] = datetime.now(timezone.utc).isoformat()
    doc["criado_por"] = usuario["email"]
    result = await db.clientes.insert_one(doc)
    return {"id": str(result.inserted_id), "referencia": doc["referencia"], "message": "Cliente cadastrado!"}


@app.put("/api/clientes/{cliente_id}")
async def atualizar_cliente(cliente_id: str, cliente: ClienteSchema, usuario=Depends(verificar_token)):
    doc = cliente.dict()
    doc["atualizado_em"] = datetime.now(timezone.utc).isoformat()
    doc["atualizado_por"] = usuario["email"]
    result = await db.clientes.update_one(
        {"_id": to_object_id(cliente_id)},
        {"$set": doc}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Cliente não encontrado")
    return {"message": "Cliente atualizado!"}


@app.delete("/api/clientes/{cliente_id}")
async def deletar_cliente(cliente_id: str, admin=Depends(apenas_admin)):
    result = await db.clientes.delete_one({"_id": to_object_id(cliente_id)})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Cliente não encontrado")
    return {"message": "Cliente removido!"}


# ============================================================
# 12. Pedidos
# ============================================================

@app.get("/api/pedidos")
async def listar_pedidos(usuario=Depends(verificar_token)):
    pedidos = await db.pedidos.find().to_list(length=500)
    return [serialize(p) for p in pedidos]


@app.get("/api/pedidos/{pedido_id}")
async def buscar_pedido(pedido_id: str, usuario=Depends(verificar_token)):
    pedido = await db.pedidos.find_one({"_id": to_object_id(pedido_id)})
    if not pedido:
        raise HTTPException(status_code=404, detail="Pedido não encontrado")
    return serialize(pedido)


@app.post("/api/pedidos", status_code=201)
async def criar_pedido(pedido: PedidoSchema, usuario=Depends(verificar_token)):
    doc = pedido.dict()
    if not doc.get("numero_oc"):
        doc["numero_oc"] = gerar_numero_oc()
    doc["criado_em"] = datetime.now(timezone.utc).isoformat()
    doc["criado_por"] = usuario["email"]
    result = await db.pedidos.insert_one(doc)
    return {"id": str(result.inserted_id), "numero_oc": doc["numero_oc"], "message": "Pedido registrado!"}


@app.put("/api/pedidos/{pedido_id}")
async def atualizar_pedido(pedido_id: str, pedido: PedidoSchema, usuario=Depends(verificar_token)):
    pedido_atual = await db.pedidos.find_one({"_id": to_object_id(pedido_id)})
    if not pedido_atual:
        raise HTTPException(status_code=404, detail="Pedido não encontrado")
    if pedido_atual.get("status") == "NF_EMITIDA" and usuario.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Pedido com NF emitida não pode ser editado")
    doc = pedido.dict()
    doc["atualizado_em"] = datetime.now(timezone.utc).isoformat()
    doc["atualizado_por"] = usuario["email"]
    await db.pedidos.update_one({"_id": to_object_id(pedido_id)}, {"$set": doc})
    return {"message": "Pedido atualizado com sucesso!"}


@app.delete("/api/pedidos/{pedido_id}")
async def deletar_pedido(pedido_id: str, admin=Depends(apenas_admin)):
    result = await db.pedidos.delete_one({"_id": to_object_id(pedido_id)})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Pedido não encontrado")
    return {"message": "Pedido removido!"}


# ============================================================
# 13. Dashboard
# ============================================================

@app.get("/api/dashboard/stats")
async def get_dashboard_stats(usuario=Depends(verificar_token)):
    now = datetime.now(timezone.utc)
    inicio_mes = now.replace(day=1).strftime("%Y-%m-%d")
    fim_mes = now.strftime("%Y-%m-%d")

    todos_pedidos = await db.pedidos.find().to_list(length=1000)

    inicio_mes_iso = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0).isoformat()
    pedidos_mes = [p for p in todos_pedidos if p.get("criado_em", "") >= inicio_mes_iso]

    total_clientes = await db.clientes.count_documents({})
    total_materiais = await db.materiais.count_documents({})

    pedidos_por_status = {}
    for p in todos_pedidos:
        s = p.get("status", "PENDENTE")
        pedidos_por_status[s] = pedidos_por_status.get(s, 0) + 1

    # Comissão prevista = pedidos com data_entrega no mês atual (independente de quando foram criados)
    pedidos_entrega_mes = [
        p for p in todos_pedidos
        if p.get("data_entrega", "")[:7] == now.strftime("%Y-%m")
        and p.get("status") not in ("CANCELADO", "NF_EMITIDA")
    ]

    comissao_prevista = round(sum(
        item.get("comissao_valor", 0)
        for p in pedidos_entrega_mes for item in p.get("itens", [])
    ), 2)

    # Tonelagem implantada = pedidos com data_entrega no mês
    tonelagem_implantada = sum(
        p.get("peso_total", 0) for p in pedidos_entrega_mes
    ) / 1000

    # Faturamento do mês = NFs com data_emissao no mês atual
    # NFs sem data_emissao são ignoradas (não contabilizadas no mês)
    notas_mes = await db.notas_fiscais.find({
        "data_emissao": {"$gte": inicio_mes, "$lte": fim_mes}
    }).to_list(length=1000)

    todas_notas_mes = notas_mes

    pedidos_ids_faturados_mes = {str(n.get("pedido_id", "")) for n in todas_notas_mes if n.get("pedido_id")}

    tonelagem_faturada = 0.0
    faturado_mes_valor = 0.0
    comissao_realizada = 0.0

    for p in todos_pedidos:
        pid_obj = str(p.get("_id", ""))
        pid_str = str(p.get("id", ""))
        if pid_obj in pedidos_ids_faturados_mes or pid_str in pedidos_ids_faturados_mes:
            tonelagem_faturada += p.get("peso_total", 0)
            faturado_mes_valor += p.get("valor_total", 0)
            for item in p.get("itens", []):
                comissao_realizada += item.get("comissao_valor", 0)

    if comissao_realizada == 0:
        comissao_realizada = sum(n.get("comissao_total", 0) for n in todas_notas_mes)

    tonelagem_faturada = tonelagem_faturada / 1000

    comissao_realizada = round(comissao_realizada, 2)
    comissao_mes = comissao_realizada
    comissoes_a_receber = round(comissao_prevista - comissao_realizada, 2)
    pedidos_mes_valor = sum(p.get("valor_total", 0) for p in pedidos_mes)

    return {
        "tonelagem_implantada": round(tonelagem_implantada, 3),
        "comissao_prevista": comissao_prevista,
        "tonelagem_faturada": round(tonelagem_faturada, 3),
        "comissao_realizada": comissao_realizada,
        "pedidos_mes_valor": round(pedidos_mes_valor, 2),
        "faturado_mes_valor": round(faturado_mes_valor, 2),
        "comissao_mes": comissao_mes,
        "comissoes_a_receber": comissoes_a_receber,
        "total_clientes": total_clientes,
        "total_materiais": total_materiais,
        "total_pedidos": len(todos_pedidos),
        "pedidos_mes": len(pedidos_mes),
        "pedidos_por_status": pedidos_por_status,
    }


# ============================================================
# 14. Metas
# ============================================================

class MetaSchema(BaseModel):
    cliente_id: str
    mes: str
    ano: int = 2026
    valor_ton: float = 0.0

@app.get("/api/metas")
async def listar_metas(usuario=Depends(verificar_token)):
    metas = await db.metas.find().to_list(length=1000)
    return [serialize(m) for m in metas]

@app.post("/api/metas", status_code=201)
async def criar_meta(meta: MetaSchema, usuario=Depends(apenas_admin)):
    existente = await db.metas.find_one({"cliente_id": meta.cliente_id, "mes": meta.mes, "ano": meta.ano})
    if existente:
        await db.metas.update_one(
            {"_id": existente["_id"]},
            {"$set": {"valor_ton": meta.valor_ton, "atualizado_em": datetime.now(timezone.utc).isoformat()}}
        )
        return {"message": "Meta atualizada!"}
    doc = meta.dict()
    doc["criado_em"] = datetime.now(timezone.utc).isoformat()
    doc["criado_por"] = usuario["email"]
    await db.metas.insert_one(doc)
    return {"message": "Meta cadastrada!"}

@app.put("/api/metas/{meta_id}")
async def atualizar_meta(meta_id: str, meta: MetaSchema, usuario=Depends(apenas_admin)):
    result = await db.metas.update_one(
        {"_id": to_object_id(meta_id)},
        {"$set": {**meta.dict(), "atualizado_em": datetime.now(timezone.utc).isoformat()}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Meta não encontrada")
    return {"message": "Meta atualizada!"}

@app.delete("/api/metas/{meta_id}")
async def deletar_meta(meta_id: str, usuario=Depends(apenas_admin)):
    result = await db.metas.delete_one({"_id": to_object_id(meta_id)})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Meta não encontrada")
    return {"message": "Meta removida!"}


# ============================================================
# 15. Health Check
# ============================================================

@app.get("/")
async def health_check():
    return {"status": "online", "sistema": "ERP Vendas 2026", "versao": "3.0"}


@app.get("/api/health")
async def api_health():
    try:
        await db.command("ping")
        return {"status": "online", "banco": "conectado"}
    except Exception:
        return {"status": "online", "banco": "erro na conexão"}


# ============================================================
# 16. Notas Fiscais & Vencimentos
# ============================================================

class NotaFiscalSchema(BaseModel):
    numero_nf: str
    data_emissao: Optional[str] = None
    pedido_id: str
    valor_total: float
    numero_parcelas: int = 1
    datas_manuais: List[str] = []
    qtde_entregue: Optional[int] = None
    motivo_entrega: Optional[str] = None

class VencimentoUpdateSchema(BaseModel):
    status: Optional[str] = None
    data_pagamento: Optional[str] = None

@app.get("/api/notas-fiscais")
async def listar_notas(usuario=Depends(verificar_token)):
    notas = await db.notas_fiscais.find().sort("criado_em", -1).to_list(length=1000)
    return [serialize(n) for n in notas]

@app.post("/api/notas-fiscais", status_code=201)
async def criar_nota(nota: NotaFiscalSchema, usuario=Depends(apenas_admin)):
    try:
        pedido = await db.pedidos.find_one({"_id": to_object_id(nota.pedido_id)})
    except:
        raise HTTPException(status_code=404, detail="Pedido não encontrado")
    if not pedido:
        raise HTTPException(status_code=404, detail="Pedido não encontrado")

    comissao_percent = 0.0
    numero_fe = pedido.get("numero_fe", "")
    if numero_fe:
        material = await db.materiais.find_one({"numero_fe": numero_fe})
        if material:
            comissao_percent = float(material.get("comissao", 0) or 0)

    if comissao_percent <= 0:
        itens = pedido.get("itens", [])
        if itens:
            comissao_percent = float(itens[0].get("comissao_percent", 0) or 0)

    comissao_total = nota.valor_total * comissao_percent / 100
    comissao_por_parcela = round(comissao_total / nota.numero_parcelas, 2)

    now = datetime.now(timezone.utc).isoformat()
    nota_doc = {
        "numero_nf": nota.numero_nf,
        "data_emissao": nota.data_emissao,
        "pedido_id": nota.pedido_id,
        "cliente_nome": pedido.get("cliente_nome", ""),
        "cliente_id": pedido.get("cliente_id", ""),
        "item_nome": pedido.get("item_nome", ""),
        "numero_fe": pedido.get("numero_fe", ""),
        "numero_fabrica": pedido.get("numero_fabrica", ""),
        "valor_total": nota.valor_total,
        "comissao_percent": comissao_percent,
        "comissao_total": comissao_total,
        "numero_parcelas": nota.numero_parcelas,
        "qtde_entregue": nota.qtde_entregue,
        "motivo_entrega": nota.motivo_entrega,
        "criado_em": now,
        "criado_por": usuario["email"],
    }
    result = await db.notas_fiscais.insert_one(nota_doc)
    nota_id = str(result.inserted_id)

    hoje = datetime.now(timezone.utc).date()
    for i in range(nota.numero_parcelas):
        if i < len(nota.datas_manuais) and nota.datas_manuais[i]:
            data_venc_str = nota.datas_manuais[i]
            data_venc = datetime.strptime(data_venc_str, "%Y-%m-%d").date()
        else:
            data_venc = hoje + timedelta(days=30 * (i + 1))
            data_venc_str = data_venc.isoformat()

        status_venc = "Pago" if data_venc < hoje else "Pendente"
        venc_doc = {
            "nota_fiscal_id": nota_id,
            "pedido_id": nota.pedido_id,
            "cliente_nome": pedido.get("cliente_nome", ""),
            "numero_nf": nota.numero_nf,
            "parcela": i + 1,
            "total_parcelas": nota.numero_parcelas,
            "data_vencimento": data_venc_str,
            "comissao_calculada": comissao_por_parcela,
            "valor_parcela": round(nota.valor_total / nota.numero_parcelas, 2),
            "status": status_venc,
            "data_pagamento": None,
            "criado_em": now,
        }
        await db.vencimentos.insert_one(venc_doc)

    await db.pedidos.update_one(
        {"_id": to_object_id(nota.pedido_id)},
        {"$set": {"status": "NF_EMITIDA", "nota_fiscal_id": nota_id}}
    )
    return {"message": "NF criada e vencimentos gerados!", "nota_id": nota_id, "vencimentos_gerados": nota.numero_parcelas}

class NotaFiscalUpdateSchema(BaseModel):
    data_emissao: Optional[str] = None
    numero_nf: Optional[str] = None

@app.put("/api/notas-fiscais/{nota_id}")
async def atualizar_nota(nota_id: str, data: NotaFiscalUpdateSchema, usuario=Depends(apenas_admin)):
    campos = {k: v for k, v in data.dict().items() if v is not None}
    if not campos:
        raise HTTPException(status_code=400, detail="Nenhum campo para atualizar")
    result = await db.notas_fiscais.update_one(
        {"_id": to_object_id(nota_id)},
        {"$set": {**campos, "atualizado_em": datetime.now(timezone.utc).isoformat()}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="NF não encontrada")
    return {"message": "NF atualizada!"}

@app.delete("/api/notas-fiscais/{nota_id}")
async def deletar_nota(nota_id: str, usuario=Depends(apenas_admin)):
    await db.vencimentos.delete_many({"nota_fiscal_id": nota_id})
    result = await db.notas_fiscais.delete_one({"_id": to_object_id(nota_id)})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="NF não encontrada")
    return {"message": "NF e vencimentos removidos!"}

@app.get("/api/vencimentos")
async def listar_vencimentos(usuario=Depends(verificar_token)):
    hoje = datetime.now(timezone.utc).date()
    hoje_iso = hoje.isoformat()

    # Lógica de comissão com defasagem de 1 mês:
    # Vencimentos do cliente em MES X → comissão paga no último dia útil de MES X+1
    # Status:
    #   Pendente  → data_vencimento ainda não chegou ou mês seguinte ainda não fechou
    #   Atrasado  → já passou o último dia do mês seguinte ao vencimento e não foi pago
    #   Pago      → confirmado manualmente

    vencimentos = await db.vencimentos.find({"status": {"$ne": "Pago"}}).to_list(length=1000)
    for v in vencimentos:
        data_venc_str = v.get("data_vencimento")
        if not data_venc_str:
            continue
        try:
            data_venc = datetime.strptime(data_venc_str, "%Y-%m-%d").date()
            # Último dia do mês seguinte ao vencimento
            mes_pagamento = data_venc.month + 1
            ano_pagamento = data_venc.year
            if mes_pagamento > 12:
                mes_pagamento = 1
                ano_pagamento += 1
            import calendar
            ultimo_dia = calendar.monthrange(ano_pagamento, mes_pagamento)[1]
            prazo_pagamento = datetime(ano_pagamento, mes_pagamento, ultimo_dia).date()

            if hoje > prazo_pagamento:
                novo_status = "Atrasado"
            else:
                novo_status = "Pendente"

            if v.get("status") != novo_status:
                await db.vencimentos.update_one(
                    {"_id": v["_id"]},
                    {"$set": {"status": novo_status}}
                )
        except Exception:
            continue

    vencimentos = await db.vencimentos.find().sort("data_vencimento", 1).to_list(length=1000)
    return [serialize(v) for v in vencimentos]

@app.post("/api/admin/reset-vencimentos-pago")
async def reset_vencimentos_pago(admin=Depends(apenas_admin)):
    """
    Reseta todos os vencimentos marcados como Pago de volta para Pendente,
    para que sejam recalculados corretamente na próxima chamada de GET /vencimentos.
    Usar apenas uma vez para corrigir status definidos automaticamente pelo sistema antigo.
    """
    result = await db.vencimentos.update_many(
        {"status": "Pago", "data_pagamento": None},
        {"$set": {"status": "Pendente"}}
    )
    return {"message": f"{result.modified_count} vencimentos resetados para Pendente!"}



async def atualizar_vencimento(venc_id: str, data: VencimentoUpdateSchema, usuario=Depends(apenas_admin)):
    update = {k: v for k, v in data.dict().items() if v is not None}
    if not update:
        raise HTTPException(status_code=400, detail="Nenhum campo para atualizar")
    result = await db.vencimentos.update_one(
        {"_id": to_object_id(venc_id)},
        {"$set": update}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Vencimento não encontrado")
    return {"message": "Vencimento atualizado!"}


# ============================================================
# 17. Exportação Excel — Comissões
# ============================================================

@app.get("/api/export/comissoes")
async def exportar_comissoes(usuario=Depends(apenas_admin)):
    vencimentos = await db.vencimentos.find().sort("data_vencimento", 1).to_list(length=5000)
    notas = await db.notas_fiscais.find().to_list(length=5000)
    notas_map = {str(n["_id"]): n for n in notas}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Comissões"

    header_fill = PatternFill("solid", fgColor="0A3D73")
    header_font = Font(bold=True, color="FFFFFF", size=9)
    header_align = Alignment(horizontal="center", vertical="center")
    pago_fill   = PatternFill("solid", fgColor="D1FAE5")
    atraso_fill = PatternFill("solid", fgColor="FEE2E2")
    border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    headers = ["NF", "Cliente", "Parcela", "Total Parcelas", "Vencimento", "Valor Parcela (R$)", "Comissão (R$)", "Status", "Data Pagamento"]
    ws.append(headers)
    for col, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = border

    for v in vencimentos:
        nota = notas_map.get(str(v.get("nota_fiscal_id", "")), {})
        status_v = v.get("status", "Pendente")
        row = [
            v.get("numero_nf") or nota.get("numero_nf", ""),
            v.get("cliente_nome") or nota.get("cliente_nome", ""),
            v.get("parcela", ""),
            v.get("total_parcelas", ""),
            v.get("data_vencimento", ""),
            v.get("valor_parcela", 0),
            v.get("comissao_calculada", 0),
            status_v,
            v.get("data_pagamento") or "",
        ]
        ws.append(row)
        row_idx = ws.max_row
        fill = pago_fill if status_v == "Pago" else (atraso_fill if status_v == "Atrasado" else None)
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = border
            cell.font = Font(size=9)
            cell.alignment = Alignment(vertical="center")
            if fill:
                cell.fill = fill
            if col in [6, 7]:
                cell.number_format = 'R$ #,##0.00'

    col_widths = [12, 30, 9, 14, 14, 20, 16, 12, 16]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 20

    ws.append([])
    ws.append(["", "", "", "", "TOTAL PAGO", "", sum(v.get("comissao_calculada", 0) for v in vencimentos if v.get("status") == "Pago")])
    ws.append(["", "", "", "", "TOTAL PENDENTE", "", sum(v.get("comissao_calculada", 0) for v in vencimentos if v.get("status") == "Pendente")])
    ws.append(["", "", "", "", "TOTAL ATRASADO", "", sum(v.get("comissao_calculada", 0) for v in vencimentos if v.get("status") == "Atrasado")])

    for r in [ws.max_row - 2, ws.max_row - 1, ws.max_row]:
        ws.cell(row=r, column=5).font = Font(bold=True, size=9)
        ws.cell(row=r, column=7).font = Font(bold=True, size=9)
        ws.cell(row=r, column=7).number_format = 'R$ #,##0.00'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    nome_arquivo = f"comissoes_{datetime.now().strftime('%d-%m-%Y')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={nome_arquivo}"}
    )


# ============================================================
# 18. Admin — Corrige cliente_id nos pedidos importados
# ============================================================

@app.post("/api/admin/fix-cliente-ids")
async def fix_cliente_ids(admin=Depends(apenas_admin)):
    clientes = await db.clientes.find().to_list(length=500)
    pedidos_sem_id = await db.pedidos.find({
        "$or": [
            {"cliente_id": {"$exists": False}},
            {"cliente_id": ""},
            {"cliente_id": None}
        ]
    }).to_list(length=1000)

    vinculados = 0
    nao_encontrados = []

    for pedido in pedidos_sem_id:
        nome_pedido = (pedido.get("cliente_nome") or "").upper().strip()
        if not nome_pedido:
            continue

        cliente_match = None
        for c in clientes:
            nome_cli = (c.get("nome") or "").upper().strip()
            if nome_cli == nome_pedido:
                cliente_match = c
                break

        if not cliente_match:
            for c in clientes:
                nome_cli = (c.get("nome") or "").upper().strip()
                partes_pedido = nome_pedido.split()
                if partes_pedido and any(p in nome_cli for p in partes_pedido[:2]):
                    cliente_match = c
                    break

        if cliente_match:
            cliente_id = str(cliente_match["_id"])
            await db.pedidos.update_one(
                {"_id": pedido["_id"]},
                {"$set": {"cliente_id": cliente_id}}
            )
            pedido_id_str = str(pedido["_id"])
            await db.notas_fiscais.update_many(
                {"pedido_id": pedido_id_str, "$or": [{"cliente_id": ""}, {"cliente_id": None}, {"cliente_id": {"$exists": False}}]},
                {"$set": {"cliente_id": cliente_id}}
            )
            vinculados += 1
        else:
            nao_encontrados.append(nome_pedido)

    return {
        "message": f"{vinculados} pedidos vinculados com sucesso!",
        "nao_encontrados": list(set(nao_encontrados)),
        "total_processados": len(pedidos_sem_id)
    }


# ============================================================
# 20. Orçamentos
# ============================================================

class ItemOrcamento(BaseModel):
    fe: Optional[str] = ""
    descricao: Optional[str] = ""
    medidas: Optional[str] = ""
    unidade: Optional[str] = "Mil"
    qtd_min: Optional[float] = 0.0
    preco_milhar: Optional[float] = 0.0
    subtotal: Optional[float] = 0.0

class OrcamentoSchema(BaseModel):
    numero_proposta: Optional[str] = ""
    data_emissao: Optional[str] = ""
    validade: Optional[str] = "20 DIAS"
    # Cliente
    cliente_id: Optional[str] = ""
    cliente_nome: Optional[str] = ""
    cliente_cnpj: Optional[str] = ""
    cliente_endereco: Optional[str] = ""
    cliente_contato: Optional[str] = ""
    cliente_telefone: Optional[str] = ""
    # Itens
    itens: Optional[List[ItemOrcamento]] = []
    # Condições
    prazo_pagamento: Optional[str] = "60 dias"
    frete: Optional[str] = "CIF"
    prazo_entrega: Optional[str] = "15 dias após confirmação"
    icms: Optional[str] = "20%"
    ipi: Optional[float] = 15.0
    # Totais
    subtotal: Optional[float] = 0.0
    valor_ipi: Optional[float] = 0.0
    valor_total: Optional[float] = 0.0
    # Status
    status: Optional[str] = "PENDENTE"  # PENDENTE | ENVIADO | APROVADO | RECUSADO
    observacoes: Optional[str] = ""

def gerar_numero_proposta() -> str:
    return f"PROP-{datetime.now().strftime('%d%m%Y-%H%M%S')}"

@app.get("/api/orcamentos")
async def listar_orcamentos(usuario=Depends(verificar_token)):
    orcamentos = await db.orcamentos.find().sort("criado_em", -1).to_list(length=500)
    return [serialize(o) for o in orcamentos]

@app.get("/api/orcamentos/{orcamento_id}")
async def buscar_orcamento(orcamento_id: str, usuario=Depends(verificar_token)):
    orc = await db.orcamentos.find_one({"_id": to_object_id(orcamento_id)})
    if not orc:
        raise HTTPException(status_code=404, detail="Orçamento não encontrado")
    return serialize(orc)

@app.post("/api/orcamentos", status_code=201)
async def criar_orcamento(orc: OrcamentoSchema, usuario=Depends(verificar_token)):
    doc = orc.dict()
    if not doc.get("numero_proposta"):
        doc["numero_proposta"] = gerar_numero_proposta()
    doc["criado_em"] = datetime.now(timezone.utc).isoformat()
    doc["criado_por"] = usuario["email"]
    result = await db.orcamentos.insert_one(doc)
    return {"id": str(result.inserted_id), "numero_proposta": doc["numero_proposta"], "message": "Orçamento criado!"}

@app.put("/api/orcamentos/{orcamento_id}")
async def atualizar_orcamento(orcamento_id: str, orc: OrcamentoSchema, usuario=Depends(verificar_token)):
    doc = orc.dict()
    doc["atualizado_em"] = datetime.now(timezone.utc).isoformat()
    doc["atualizado_por"] = usuario["email"]
    result = await db.orcamentos.update_one(
        {"_id": to_object_id(orcamento_id)},
        {"$set": doc}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Orçamento não encontrado")
    return {"message": "Orçamento atualizado!"}

@app.patch("/api/orcamentos/{orcamento_id}/status")
async def atualizar_status_orcamento(orcamento_id: str, data: dict, usuario=Depends(verificar_token)):
    status_validos = ["PENDENTE", "ENVIADO", "APROVADO", "RECUSADO"]
    novo_status = data.get("status", "").upper()
    if novo_status not in status_validos:
        raise HTTPException(status_code=400, detail=f"Status inválido. Use: {status_validos}")
    await db.orcamentos.update_one(
        {"_id": to_object_id(orcamento_id)},
        {"$set": {"status": novo_status, "atualizado_em": datetime.now(timezone.utc).isoformat()}}
    )
    return {"message": f"Status atualizado para {novo_status}"}

@app.delete("/api/orcamentos/{orcamento_id}")
async def deletar_orcamento(orcamento_id: str, usuario=Depends(verificar_token)):
    result = await db.orcamentos.delete_one({"_id": to_object_id(orcamento_id)})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Orçamento não encontrado")
    return {"message": "Orçamento removido!"}


# ============================================================
# 21. Shutdown
# ============================================================

@app.on_event("shutdown")
async def shutdown():
    client.close()
    logger.info("🔴 Conexão com MongoDB encerrada")


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)